"""
Import Monday.com Excel export into structured project rows.

Reads the '1. current projects overview' sheet, skips title/header rows,
detects and skips subitem rows and the "Completed Projects" divider.
"""

import re
import openpyxl
from . import config, utils


def _parse_money(value):
    """Normalize a budget cell value to a dollar string."""
    if value is None:
        return "$TBC"
    if isinstance(value, (int, float)):
        if value == 0:
            return "$TBC"
        return f"${value:,.0f}"
    s = str(value).strip()
    if not s or s.lower() in ("tbc", "tbd", "n/a", ""):
        return "$TBC"
    return s if s.startswith("$") else f"${s}"


def _parse_date(value):
    """Normalize a date cell to (display_str, iso_str)."""
    if value is None:
        return "TBC"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.lower() in ("tbc", "tbd", "n/a"):
        return "TBC"
    return s


def _parse_list(value):
    """Split a comma/newline-delimited cell into a list of strings."""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    items = re.split(r"[,\n]+", s)
    return [i.strip() for i in items if i.strip()]


def _cell_str(value):
    """Convert cell to stripped string or empty string."""
    if value is None:
        return ""
    return str(value).strip()


def import_excel(filepath):
    """Import a Monday.com Excel export.

    Returns a list of ExcelProjectRow dicts (main project rows only,
    subitems are skipped).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find the sheet
    sheet_name = config.EXCEL_SHEET_NAME
    if sheet_name not in wb.sheetnames:
        # Try case-insensitive match
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                sheet_name = name
                break
        else:
            raise ValueError(
                f"Sheet '{config.EXCEL_SHEET_NAME}' not found. "
                f"Available: {wb.sheetnames}"
            )

    ws = wb[sheet_name]
    rows = []

    for row_idx in range(config.EXCEL_DATA_START_ROW, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        col_a_str = _cell_str(col_a)

        # Stop at the "Completed Projects" divider
        if col_a_str.lower().startswith(config.EXCEL_COMPLETED_DIVIDER.lower()):
            break

        # Skip empty rows and subitem rows (col A empty or "Subitems")
        if not col_a_str or col_a_str == "Subitems":
            continue

        # Skip header echo rows
        if col_a_str == "Name":
            continue

        # Parse the name
        client, project, qualifier = utils.parse_excel_name(col_a_str)

        # Read all mapped columns
        raw = {}
        for col_idx, field_name in config.EXCEL_COLUMN_MAP.items():
            raw[field_name] = ws.cell(row=row_idx, column=col_idx).value

        # Map phase
        raw_phase = _cell_str(raw.get("project_phase"))
        canonical_phase = config.EXCEL_PHASE_MAP.get(raw_phase, "TBC")

        row = {
            "excel_row": row_idx,
            "raw_name": col_a_str,
            "client": client,
            "project": project,
            "qualifier": qualifier,
            "bfa_lead": _cell_str(raw.get("bfa_lead")),
            "project_state": _cell_str(raw.get("project_state")) or "TBC",
            "project_phase": canonical_phase,
            "municipality": _cell_str(raw.get("municipality")),
            "address": _cell_str(raw.get("address")),
            "selected_artist": _cell_str(raw.get("selected_artist")),
            "shortlisted_artists": _parse_list(raw.get("shortlisted_artists")),
            "artwork_title": _cell_str(raw.get("artwork_title")),
            "fabricator": _cell_str(raw.get("fabricator")),
            "architect": _cell_str(raw.get("architect")),
            "landscape_architect": _cell_str(raw.get("landscape_architect")),
            "developer_contact": _cell_str(raw.get("developer_contact")),
            "selection_panel": _parse_list(raw.get("selection_panel")),
            "community_advisors": _parse_list(raw.get("community_advisors")),
            "billing_entity": _cell_str(raw.get("billing_entity")),
            "selection_process": _cell_str(raw.get("selection_process")),
            "project_budget": _parse_money(raw.get("project_budget")),
            "artist_fee": _parse_money(raw.get("artist_fee")),
            "consultant_fee": _parse_money(raw.get("consultant_fee")),
            "target_completion": _cell_str(raw.get("target_completion")) or "TBC",
            "target_completion_start": _parse_date(raw.get("target_completion_start")),
            "building_occupancy": _cell_str(raw.get("building_occupancy")),
            "milestone_sp1": _parse_date(raw.get("milestone_sp1")),
            "milestone_sp2": _parse_date(raw.get("milestone_sp2")),
            "dp_issuance_date": _parse_date(raw.get("dp_issuance_date")),
            "artist_orientation": _parse_date(raw.get("artist_orientation")),
            "project_kickoff": _parse_date(raw.get("project_kickoff")),
            "artist_contract_signed": _parse_date(raw.get("artist_contract_signed")),
        }

        rows.append(row)

    return rows
